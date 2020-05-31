import datetime
import loadConfig
from time import perf_counter

from openpyxl import load_workbook, Workbook


def exec_time(func):
    def make_decorater(*args, **kwargs):  # 接受调用语句的实参，在下面传递给被装饰函数（原函数）
        print("开始解析【%s】表格" % args[1])
        start = perf_counter()
        func(*args, **kwargs)  # 如果在这里return，则下面的代码无法执行，所以引用并在下面返回
        end = perf_counter()
        print(end - start)

    return make_decorater


shop_map = {}
shop_map_key = set()


# 获取配置信息
config_obj = loadConfig.get_config()


def open_excel(path):
    print("--------【{}】打开Excel--------".format(datetime.datetime.now()))
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    print("--------【{}】打开ExcelEnd--------".format(datetime.datetime.now()))
    # 创建一个新的excel
    write_book = Workbook()
    # 设置第一个sheet页名字为汇总
    write_book.active.title = "汇总"
    sheet_num = 1

    for sheet_name in wb.sheetnames:
        parse_sheet(wb, sheet_name)
        if len(shop_map_key) > 0:
            save_excel(write_book, sheet_num, sheet_name)
            # 清空缓存
            shop_map.clear()
            shop_map_key.clear()
        sheet_num = sheet_num + 1


def save_excel(write_book, sheet_num, sheet_name):
    n_sheet = write_book.create_sheet(sheet_name)
    i = 1
    for group_column in config_obj.all_columns:
        """ 遍历所有列，设置表头 """
        n_sheet.cell(1, i, group_column)
        i = i + 1

    shop_name = None
    shop_count = 0
    money_sum = 0
    n_sheet_row = 1
    # 将缓存数据写入Excel
    for shop_key in shop_map_key:
        if shop_key in shop_map:
            # 默认都从第二行开始
            n_sheet_row = n_sheet_row + 1
            info_map = shop_map[shop_key]
            col_index = 1
            for count_column in config_obj.all_columns:
                info_val = info_map[count_column]
                n_sheet.cell(n_sheet_row, col_index, info_val)
                col_index = col_index + 1
            # 汇总
            # shop_name = info.shop_name
            # shop_count = shop_count + info.shop_count
            # money_sum = money_sum + info.money_sum
        else:
            print(shop_key)

    # count_sheet = write_book.get_sheet_by_name("汇总")
    # count_sheet.cell(sheet_num, 1, shop_name)
    # count_sheet.cell(sheet_num, 2, shop_count)
    # count_sheet.cell(sheet_num, 3, money_sum)
    write_book.save("C:\\Users\\l2503\\Desktop\\汇总.xlsx")


@exec_time
def parse_sheet(wb, sheet_name):
    # print(sheet_name)
    sheet_obj = wb[sheet_name]
    # 返回解析对象
    parse_info = get_parse_info(sheet_obj)
    if parse_info is not None:
        parse_core(sheet_obj, parse_info)

    return


def get_parse_info(sheet_obj):
    """ 返回需要解析的信息 """
    parse_info = ParseInfo()
    # 行序号
    row_num = 1
    for row in sheet_obj.iter_rows(min_row=row_num, max_row=3, values_only=True):
        print("--------【{}】开始解析{:d}行--------".format(datetime.datetime.now(), row_num))
        col_idx = 0
        for cell in row:
            # 列序号
            col_idx = col_idx + 1
            # print(isinstance(cell.value, str))
            if cell is None or isinstance(cell, str) is False:
                continue
            for group_column in config_obj.all_columns:
                if verify_title(cell, group_column) is True:
                    parse_info.cell_map[group_column] = col_idx
            # print(str(parse_info.cell_map))
            # 如果两边长度一致
            if len(parse_info.cell_map) == len(config_obj.all_columns):
                parse_info.min_row_num = row_num
                # 返回需要解析的信息
                return parse_info
        row_num = row_num + 1
    return None


def parse_core(sheet_obj, parse_info):
    """ 解析核心逻辑 """
    row_index = parse_info.min_row_num + 1
    print("核心逻辑--------总行数：【{}】，【{}】从第【{}】开始解析--------".
          format(sheet_obj.max_row, datetime.datetime.now(), row_index))
    break_count = 0
    row_num = 0
    # 采用只读模式，通过对tup进行处理
    for row in sheet_obj.iter_rows(min_row=row_index, values_only=True):
        # print(row)
        # 空行>10 则代表后面都为无效表格，进行跳出
        if break_count > 10:
            break

        key_str = ""
        none_column_count = 0
        info_map = {}
        for group_column in config_obj.group_columns:
            """ 遍历分组列 """
            group_index = parse_info.cell_map[group_column]-1
            title_str = str(row[group_index])
            # 赋值
            info_map[group_column] = title_str

            if title_str == 'None':
                none_column_count = none_column_count + 1
            else:
                key_str = key_str + "," + title_str

        if none_column_count == len(config_obj.group_columns):
            # 如果字段都为空，则代表空行 ，空行加一
            # print(sheet_obj, row_num, shop_column, user_column, money_column, count_column)
            break_count = break_count + 1
            continue

        for count_column in config_obj.count_columns:
            """ 遍历汇总列 """
            count_index = parse_info.cell_map[count_column]-1
            count_val = row[count_index]
            if count_val is None or isinstance(count_val, str) is True:
                count_val = 0.0
            info_map[count_column] = count_val



        if key_str in shop_map:
            cache_map = shop_map[key_str]
            for count_column in config_obj.count_columns:
                cache_map[count_column] = cache_map[count_column] + info_map[count_column]
        else:
            shop_map_key.add(key_str)
            shop_map[key_str] = info_map

        row_num = row_num + 1

    print("核心逻辑--------【{}】已解析{:d}行--------".format(datetime.datetime.now(), row_num))


def verify_title(title, key_str):
    """验证表头是否包含关键字"""
    return title.find(key_str) != -1


class ParseInfo:
    """ 解析信息 """
    def __init__(self):
        self.min_row_num = 0
        self.cell_map = {}


# class ShopInfo:
#     """ 商品信息 """
#
#     def __init__(self, shop_name, user_name, shop_count, money_sum):
#         self.shop_name = shop_name
#         self.user_name = user_name
#         self.shop_count = shop_count
#         self.money_sum = money_sum
#
#     def add_count(self, count):
#         try:
#             self.shop_count = self.shop_count + count
#         except TypeError as e:
#             print(type(count))
#
#     def add_money(self, money):
#         self.money_sum = self.money_sum + money
#
#     def __str__(self):
#         return self.shop_name + "--" + self.user_name + "--" + str(self.shop_count) + "--" + str(self.money_sum)


# def write_excel():


if __name__ == '__main__':
    open_excel("C:\\Users\\l2503\\Desktop\\线上不开票客户4月份销售明细.xlsx")
